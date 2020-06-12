# coding = utf-8

import time
import openpyxl
from openpyxl.styles import Border, Side, Font

class ParseExcel(object):

    def __init__(self):
        self.workbook = None
        self.excelFile = None
        self.font = Font(color=None) #设置字体颜色
        #颜色对应的RGB值
        self.RGBDict = {'red': 'FFFF3030', 'green': 'FF008B00',"black":"000000"}

    def loadWorkBook(self,excelPathAndName):
        #将Excel文件加载到内存，并获取workbook对象
        try:
            self.workbook = openpyxl.load_workbook(excelPathAndName)
        except Exception as e:
            raise e
        self.excelFile = excelPathAndName

    def getSheetByName(self,sheetNmae):
        #根据sheet名称获取该sheet对象
        try:
            sheet =self.workbook.get_sheet_by_name(sheetNmae)
            return sheet
        except Exception as e:
            raise e

    def getsheetByIndex(self,sheetIndex):
        #根据sheet的索引号获取该sheet对象
        try:
            sheetname = self.workbook.get_sheet_names()[sheetIndex]
        except Exception as e:
            raise e
        sheet = self.workbook.get_sheet_by_name(sheetname)
        return sheet

    def getRowsNumber(self,sheet):
        #获取sheet中有数据区域的结束行号
        return sheet.max_row

    def getColsNumber(self_,sheet):
        #获取sheet中有数据区域的结束列号
        return sheet.max_column

    def getStartRowNumber(self,sheet):
        #获取有数区域 的开始行号
        return sheet.min_row
    def getStartColNumber(self,sheet):
        #获取有数据区域的开始列号
        return sheet.min_column

    def getRow(self,sheet,rowNo):
        #获取sheet中某一行，返回的是这一行所有的数据内容组成tuple
        #下标从1开始，sheetrows[1]表示第一行
        try:
            return list(sheet.rows)[rowNo - 1]
        except Exception as e:
            raise e

    def getColumn(self,sheet,colNo):

         #获取sheet中某一列，返回的是这一列所有数据组成的 tuple
         #下标从1开始，sheetcolumn[1]表示第一列
        try:
          return list(sheet.columns)[colNo - 1]
        except Exception as e:
           raise e

    def getCellOfValue(self,sheet,coordinate=None,rowNo=None,colNo =None):
        #根据单元格所在的位置索引获取该单元格中的值，下标从1开始，
        #sheet.cell(row =1,column =1).value表示EXcel中第一行第一列的值
        if coordinate!= None:
            try:
                return sheet[coordinate].value
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not  None:
            try:
                return sheet.cell(row =rowNo,column=colNo).value
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficent Coordinate of cell!")

    def getCellOfObject(self,sheet,coordinate = None,rowNo=None,colsNo =None):
        #获取某个单元格的对象，可以根据单元格所在的位置的数字索引
        #也可以直接根据Excel中单元格的编码及坐标，如getCellOfObject(sheet,coordinate = 'A1') 或者
        #如getCellOfObject(sheet,roNo =1 ,colsNo = 2)
        if coordinate != None:
            try:
                return sheet.cell(coordinate=coordinate)
            except Exception as e:
                raise e
        elif coordinate == None and rowNo is not None and colsNo is not None:
            try:
                return sheet.cell(row= rowNo,column = colsNo)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficent Coordinate of cell!")


    def writeCell(self, sheet, content, coordinate=None, rowNo=None, colNo=None, style=None):

        # 根据单元格在Excel中的编码坐标或者数字索引坐标向单元格中写入数据，下标从1开始
        # 参数style表示字体的颜色的名字，如red，green
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = content
                if style is not None:
                    sheet.cell(coordinate=coordinate).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)

            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = content
                if style is not None:
                    sheet.cell(row=rowNo, column=colNo).font = Font(color=self.RGBDict[style])
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        else:
            raise Exception("Insufficient Coordinates of cell!")


    def writeCellCurrentTime(self, sheet, coordinate=None, rowNo=None, colNo=None):
        # 写入当前时间，下标从1开始
        now = int(time.time())  # 显示为时间戳
        timeArray = time.localtime(now)
        currentTime = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
        if coordinate is not None:
            try:
                sheet.cell(coordinate=coordinate).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e
        elif coordinate is None and rowNo is not None and colNo is not None:
            try:
                sheet.cell(row=rowNo, column=colNo).value = currentTime
                self.workbook.save(self.excelFile)
            except Exception as e:
                raise e


if __name__ =='__main__':
    pe = ParseExcel()
    # 测试所用的excel文件“126邮箱联系人.xlsx”请自行创建
    pe.loadWorkBook(r'C:\Users\Avidly\Desktop\接口测试\InterfaceTest\unitest\common\126邮箱联系人.xlsx')

    print(pe.getSheetByName(u"126账号联系人"))
    print(pe.getsheetByIndex(0))










